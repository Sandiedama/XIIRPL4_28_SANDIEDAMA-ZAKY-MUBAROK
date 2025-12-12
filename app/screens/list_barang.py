from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLineEdit, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QDialog, QLabel, QFormLayout, QSpinBox, QComboBox, QFileDialog, QDialogButtonBox
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QPixmap, QImage
import os
from database.db import get_connection
from utils.helpers import resource_path, output_assets_path, get_current_user, log_aktivitas, is_admin
from utils.theme import theme
import openpyxl
import traceback
import webbrowser
from PyQt6.QtWidgets import QWidget, QVBoxLayout, QLabel, QPushButton

## Barcode preview/generate dialogs removed

class ListBarangScreen(QWidget):
    data_changed = pyqtSignal()
    def __init__(self, parent=None, floating_mode=False):
        super().__init__(parent)
        self.floating_mode = floating_mode
        self.init_ui()
        self.load_data()
        if self.floating_mode:
            self.hide_non_preview_buttons()

    def hide_non_preview_buttons(self):
        if hasattr(self, 'btn_add'):
            self.btn_add.hide()
        if hasattr(self, 'btn_edit'):
            self.btn_edit.hide()
        if hasattr(self, 'btn_delete'):
            self.btn_delete.hide()
        if hasattr(self, 'btn_delete_all'):
            self.btn_delete_all.hide()
        if hasattr(self, 'btn_generate_all'):
            self.btn_generate_all.hide()
        if hasattr(self, 'btn_generate_barcode'):
            self.btn_generate_barcode.hide()
        if hasattr(self, 'btn_set_rop'):
            self.btn_set_rop.hide()
        # Jika ada tombol import, dsb, hide juga
        for child in self.findChildren(QPushButton):
            if child.text().lower().startswith('import'):
                child.hide()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(16)
        # Header dengan tombol floating window (hanya jika bukan floating)
        header_row = QHBoxLayout()
        header = QLabel('List Barang')
        self.header = header
        header.setAlignment(Qt.AlignmentFlag.AlignLeft)
        header_row.addWidget(header)
        header_row.addStretch()
        if not getattr(self, 'floating_mode', False):
            floating_btn = QPushButton()
            floating_btn.setToolTip('Buka Floating Window')
            floating_btn.setFixedSize(36, 36)
            floating_btn.setStyleSheet('''
                QPushButton {
                    background: #FFF3F0;
                    color: #FF2800;
                    border: 2px solid #FF2800;
                    border-radius: 18px;
                    font-size: 20px;
                    font-weight: bold;
                }
                QPushButton:hover { background: #FF2800; color: #FFF; }
            ''')
            floating_btn.setText('🪟')
            floating_btn.clicked.connect(self.open_in_floating_window)
            header_row.addWidget(floating_btn)
        layout.addLayout(header_row)
        # Card statistik jumlah barang
        stat_row = QHBoxLayout()
        stat_card = QLabel()
        stat_card.setObjectName('statCard')
        # Stat card styling will be applied in apply_theme method
        stat_card.setText('Total Barang: ...')
        stat_row.addWidget(stat_card)
        stat_row.addStretch()
        layout.addLayout(stat_row)
        self.stat_card = stat_card
        try:
            from PyQt6.QtWidgets import QGraphicsDropShadowEffect
            shadow = QGraphicsDropShadowEffect()
            shadow.setBlurRadius(18)
            shadow.setOffset(0, 6)
            stat_card.setGraphicsEffect(shadow)
        except Exception:
            pass
        # Tombol utama horizontal (responsive)
        btn_row = QHBoxLayout()
        btn_row.setSpacing(14)
        
        # Tombol Refresh (untuk semua user)
        self.btn_refresh = QPushButton('Refresh')
        self.btn_refresh.setStyleSheet('''
            QPushButton { background:#FF2800; color:#FFF; font-size:16px; padding:10px 20px; border-radius:10px; font-weight:600; }
            QPushButton:hover { background:#e02400; }
        ''')
        self.btn_refresh.setMinimumWidth(120)
        self.btn_refresh.clicked.connect(self.refresh_data)
        btn_row.addWidget(self.btn_refresh)
        
        # Tombol Tambah untuk semua user
        self.btn_add = QPushButton('Tambah')
        self.btn_add.setStyleSheet('''
            QPushButton { background:#FF2800; color:#FFF; font-size:16px; padding:10px 20px; border-radius:10px; font-weight:600; }
            QPushButton:hover { background:#e02400; }
        ''')
        self.btn_add.setMinimumWidth(120)
        self.btn_add.clicked.connect(self.add_barang)
        btn_row.addWidget(self.btn_add)

        # Tombol Export PDF List Barang
        self.btn_export = QPushButton('Export List (PDF)')
        self.btn_export.setStyleSheet('''
            QPushButton { background:#FF2800; color:#FFF; font-size:16px; padding:10px 20px; border-radius:10px; font-weight:600; }
            QPushButton:hover { background:#e02400; }
        ''')
        self.btn_export.setMinimumWidth(140)
        self.btn_export.clicked.connect(self.export_list_barang_pdf)
        btn_row.addWidget(self.btn_export)
        
        # Tombol yang hanya untuk admin
        if is_admin():
            btn_import = QPushButton('Import Excel')
            btn_import.setStyleSheet('''
                QPushButton { background:#FF2800; color:#FFF; font-size:16px; padding:10px 20px; border-radius:10px; font-weight:600; }
                QPushButton:hover { background:#e02400; }
            ''')
            btn_import.setMinimumWidth(120)
            btn_import.clicked.connect(self.import_excel)
            btn_row.addWidget(btn_import)

        # Hapus fitur barcode button
        
        btn_row.addStretch()
        layout.addLayout(btn_row)
        # Pencarian, Filter, dan Control Bar
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText('Cari Part Number / Nama Barang / Kategori...')
        self.search_input.setStyleSheet('''
            QLineEdit { font-size:15px; padding:10px 12px; border-radius:10px; border:1px solid #EEF1F4; }
            QLineEdit:focus { border:1px solid #FF2800; }
        ''')
        self.search_input.textChanged.connect(self.search)
        search_layout.addWidget(self.search_input)
        
        # Filter kategori
        search_layout.addWidget(QLabel('Filter Kategori:'))
        self.filter_kategori = QComboBox()
        self.filter_kategori.addItem('Semua Kategori')
        # Import kategori dari utils
        from utils.categories import get_common_categories
        categories = get_common_categories()
        for category in categories:
            if category:  # Skip empty category
                self.filter_kategori.addItem(category)
        self.filter_kategori.setStyleSheet('font-size:15px; padding:8px 10px; border-radius:10px; border:1px solid #EEF1F4;')
        self.filter_kategori.currentTextChanged.connect(self.filter_by_category)
        search_layout.addWidget(self.filter_kategori)

        # Spacer kanan
        search_layout.addStretch()
        
        layout.addLayout(search_layout)
        # Tabel (responsive)
        table_container = QVBoxLayout()
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels(['Part Number', 'Nama Barang', 'Satuan', 'Stok', 'ROP'])
        header = self.table.horizontalHeader()
        if header is not None:
            header.setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            header.setFixedHeight(44)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setStyleSheet('''
            QTableWidget { font-size:15px; background:#FFF; border-radius:14px; border:1px solid #EEF1F4; gridline-color:#F0F3F6; }
            QTableWidget::item { padding:8px; }
            QTableWidget::item:selected { background:#FFE3D6; color:#222; }
            QHeaderView::section { background:#FFF; color:#FF2800; font-size:15px; font-weight:600; height:40px; border:1px solid #EEF1F4; border-radius:8px; }
        ''')
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setMinimumHeight(320)
        try:
            self.table.setShowGrid(False)
            self.table.setWordWrap(False)
        except Exception:
            pass
        table_container.addWidget(self.table)
        # Row height
        self.table.verticalHeader().setDefaultSectionSize(40)
        self.table.verticalHeader().setMinimumSectionSize(36)
        try:
            from PyQt6.QtWidgets import QGraphicsDropShadowEffect
            tshadow = QGraphicsDropShadowEffect()
            tshadow.setBlurRadius(14)
            tshadow.setOffset(0, 4)
            self.table.setGraphicsEffect(tshadow)
        except Exception:
            pass
        layout.addLayout(table_container)
        # Tombol Edit/Hapus/Barcode horizontal (responsive)
        crud_row = QHBoxLayout()
        crud_row.setSpacing(14)
        
        # Tombol yang hanya untuk admin
        if is_admin():
            self.btn_edit = QPushButton('Edit')
            self.btn_edit.setStyleSheet('''
                QPushButton { background:#FF2800; color:#FFF; font-size:16px; padding:10px 20px; border-radius:10px; font-weight:600; }
                QPushButton:hover { background:#e02400; }
            ''')
            self.btn_edit.setMinimumWidth(120)
            self.btn_edit.clicked.connect(self.edit_barang)
            crud_row.addWidget(self.btn_edit)
            
            self.btn_delete = QPushButton('Hapus')
            self.btn_delete.setStyleSheet('''
                QPushButton { background:#FF2800; color:#FFF; font-size:16px; padding:10px 20px; border-radius:10px; font-weight:600; }
                QPushButton:hover { background:#e02400; }
            ''')
            self.btn_delete.setMinimumWidth(120)
            self.btn_delete.clicked.connect(self.delete_barang)
            crud_row.addWidget(self.btn_delete)
            
            # Fitur generate barcode dihapus
            
            self.btn_delete_all = QPushButton('Hapus Semua')
            self.btn_delete_all.setStyleSheet('''
                QPushButton { background:#FFF; color:#FF2800; font-size:16px; padding:10px 20px; border:2px solid #FF2800; border-radius:10px; font-weight:600; }
                QPushButton:hover { background:#FFF3F0; }
            ''')
            self.btn_delete_all.setMinimumWidth(120)
            self.btn_delete_all.clicked.connect(self.delete_all_barang)
            crud_row.addWidget(self.btn_delete_all)
        
        
        crud_row.addStretch()
        layout.addLayout(crud_row)
        # Setelah layout tabel
        # Hapus tombol Simpan besar di bawah tabel
        # (hapus self.btn_save dan layout.addWidget(self.btn_save))
        self.setLayout(layout)
        # Apply theme
        self.apply_theme()

    def apply_theme(self):
        """Apply current theme to list barang screen"""
        # Header styling
        self.header.setStyleSheet(f'font-size:32px; color:{theme.get_color("accent")}; font-weight:bold; margin-bottom:12px;')
        
        # Stat card styling
        self.stat_card.setStyleSheet(f'background:{theme.get_color("card_background")}; color:{theme.get_color("accent")}; border-radius:16px; font-size:18px; padding:14px 28px; font-weight:bold; border:1px solid {theme.get_color("border")};')
        
        # Button styling
        button_style = theme.get_button_style('primary')
        self.btn_refresh.setStyleSheet(button_style)
        self.btn_add.setStyleSheet(button_style)
        if hasattr(self, 'btn_export'):
            self.btn_export.setStyleSheet(button_style)
        
        # Admin buttons
        if hasattr(self, 'btn_edit'):
            self.btn_edit.setStyleSheet(button_style)
        if hasattr(self, 'btn_delete'):
            self.btn_delete.setStyleSheet(button_style)
        if hasattr(self, 'btn_delete_all'):
            self.btn_delete_all.setStyleSheet(f'''
                QPushButton {{ background:{theme.get_color("card_background")}; color:{theme.get_color("accent")}; font-size:16px; padding:10px 20px; border:2px solid {theme.get_color("accent")}; border-radius:10px; font-weight:600; }}
                QPushButton:hover {{ background:{theme.get_color("table_selection")}; }}
            ''')
        
        # Input styling
        input_style = theme.get_input_style()
        self.search_input.setStyleSheet(input_style)
        self.filter_kategori.setStyleSheet(input_style)
        
        # Table styling
        table_style = theme.get_table_style()
        self.table.setStyleSheet(table_style)
        
        # Main background
        self.setStyleSheet(f'background: {theme.get_color("background")}; font-family: Segoe UI, Arial, sans-serif;')

    def load_data(self, keyword=None):
        conn = get_connection()
        c = conn.cursor()
        if keyword:
            c.execute("SELECT * FROM barang WHERE partnumber LIKE ? OR nama LIKE ? OR kategori LIKE ? ORDER BY id DESC", (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))
        else:
            c.execute("SELECT * FROM barang ORDER BY id DESC")
        rows = c.fetchall()
        self.populate_table(rows)
        conn.close()

    def search(self):
        keyword = self.search_input.text().strip()
        self.load_data(keyword if keyword else None)

    def refresh_data(self):
        self.load_data()
    
    def filter_by_category(self):
        selected_category = self.filter_kategori.currentText()
        if selected_category == 'Semua Kategori':
            self.load_data()
        else:
            self.load_data_by_category(selected_category)
    
    def load_data_by_category(self, category):
        conn = get_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM barang WHERE kategori = ? ORDER BY id DESC", (category,))
        rows = c.fetchall()
        self.populate_table(rows)
        conn.close()
    
    def populate_table(self, rows):
        """Helper method untuk mengisi tabel dengan data"""
        self.table.setRowCount(len(rows))
        for i, row in enumerate(rows):
            self.table.setItem(i, 0, QTableWidgetItem(str(row[1]) if row[1] else ''))   # Part Number
            self.table.setItem(i, 1, QTableWidgetItem(str(row[2]) if row[2] else ''))   # Nama
            self.table.setItem(i, 2, QTableWidgetItem(str(row[7]) if len(row) > 7 and row[7] else ''))  # Satuan

            
            # Stok di index 4
            stok_val = row[4]
            if stok_val is None:
                stok_str = '0'
            elif hasattr(stok_val, 'isoformat'):
                stok_str = '0'
            else:
                try:
                    stok_str = str(int(stok_val))
                except Exception:
                    stok_str = '0'
            stok_item = QTableWidgetItem(stok_str)
            self.table.setItem(i, 3, stok_item)
            
            # ROP di index 6
            rop_val = row[6] if len(row) > 6 else 0
            if rop_val is None:
                rop_str = '0'
            else:
                try:
                    rop_str = str(int(rop_val))
                except Exception:
                    rop_str = '0'
            rop_item = QTableWidgetItem(rop_str)
            self.table.setItem(i, 4, rop_item)
            
            # Warning visual untuk stok <= ROP
            try:
                stok_int = int(stok_str)
                rop_int = int(rop_str)
                if rop_int > 0 and stok_int <= rop_int:
                    # Set background merah untuk warning
                    stok_item.setBackground(Qt.GlobalColor.red)
                    stok_item.setForeground(Qt.GlobalColor.white)
                    rop_item.setBackground(Qt.GlobalColor.red)
                    rop_item.setForeground(Qt.GlobalColor.white)
                    # Tambah tooltip
                    stok_item.setToolTip(f"⚠️ WARNING: Stok ({stok_int}) sudah mencapai atau di bawah ROP ({rop_int})")
                    rop_item.setToolTip(f"⚠️ WARNING: Stok ({stok_int}) sudah mencapai atau di bawah ROP ({rop_int})")
            except Exception:
                pass
        
        # Update card statistik
        self.stat_card.setText(f'Total Barang: {len(rows)}')

    def add_barang(self):
        dialog = BarangDialog(self)
        if dialog.exec():
            data = dialog.get_data()
            conn = get_connection()
            c = conn.cursor()
            c.execute(
                "INSERT INTO barang (partnumber, nama, kategori, stok, rop, satuan) VALUES (?, ?, ?, ?, ?, ?)",
                (data['partnumber'], data['nama'], data['kategori'], data['stok'], data['rop'], data['satuan'])
            )
            barang_id = c.lastrowid
            conn.commit()
            conn.close()

            user = get_current_user()
            if user:
                log_aktivitas(user['id'], 'tambah', 'barang', barang_id, f"Tambah barang: {data['partnumber']}")
            self.load_data()
            try:
                self.data_changed.emit()
            except Exception:
                pass



    def edit_barang(self):
        if not is_admin():
            QMessageBox.warning(self, 'Akses Ditolak', 'Hanya admin yang dapat mengedit barang.')
            return

        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Error', 'Pilih barang terlebih dahulu!')
            return

        partnumber_item = self.table.item(selected, 0)  # kolom 0 = Part Number
        if not partnumber_item:
            QMessageBox.warning(self, 'Error', 'Data tidak ditemukan!')
            return

        partnumber = partnumber_item.text()

        conn = get_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM barang WHERE partnumber=?", (partnumber,))
        row = c.fetchone()
        conn.close()

        if not row:
            QMessageBox.warning(self, 'Error', 'Data tidak ditemukan!')
            return

        dialog = BarangDialog(self, data=row)
        if dialog.exec():
            data = dialog.get_data()
            conn = get_connection()
            c = conn.cursor()

            # Update data tanpa mengubah stok (stok hanya bisa diubah via menu IN/OUT)
            # Pertahankan stok yang ada di database (row[4] adalah stok)
            stok_lama = row[4] if len(row) > 4 else 0
            c.execute(
                "UPDATE barang SET partnumber=?, nama=?, kategori=?, stok=?, rop=?, satuan=? WHERE id=?",
                (data['partnumber'], data['nama'], data['kategori'], stok_lama, data['rop'], data['satuan'], row[0])
            )

            conn.commit()
            conn.close()

            user = get_current_user()
            if user:
                log_aktivitas(user['id'], 'edit', 'barang', row[0], f"Edit barang: {data['partnumber']}")
            self.load_data()
            try:
                self.data_changed.emit()
            except Exception:
                pass


    def delete_barang(self):
        if not is_admin():
            QMessageBox.warning(self, 'Akses Ditolak', 'Hanya admin yang dapat menghapus barang.')
            return
            
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Pilih Data', 'Pilih baris yang akan dihapus!')
            return
        partnumber_item = self.table.item(selected, 0)
        partnumber = partnumber_item.text() if partnumber_item is not None else ''
        reply = QMessageBox.question(self, 'Konfirmasi', f'Yakin hapus barang dengan Part Number {partnumber}?',
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            conn = get_connection()
            c = conn.cursor()
            c.execute("SELECT id FROM barang WHERE partnumber=?", (partnumber,))
            row = c.fetchone()
            barang_id = row[0] if row else None
            c.execute("DELETE FROM barang WHERE partnumber=?", (partnumber,))
            conn.commit()
            conn.close()
            user = get_current_user()
            if user and barang_id:
                log_aktivitas(user['id'], 'hapus', 'barang', barang_id, f"Hapus barang: {partnumber}")
            self.load_data()
            try:
                self.data_changed.emit()
            except Exception:
                pass

    def delete_all_barang(self):
        if not is_admin():
            QMessageBox.warning(self, 'Akses Ditolak', 'Hanya admin yang dapat menghapus semua barang.')
            return
            
        reply = QMessageBox.question(self, 'Konfirmasi', 'Yakin hapus SEMUA barang?', QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            conn = get_connection()
            c = conn.cursor()
            c.execute("DELETE FROM barang")
            conn.commit()
            conn.close()
            self.load_data()
            QMessageBox.information(self, 'Hapus Semua', 'Semua barang berhasil dihapus.')
            try:
                self.data_changed.emit()
            except Exception:
                pass

    def import_excel(self):
        if not is_admin():
            QMessageBox.warning(self, 'Akses Ditolak', 'Hanya admin yang dapat mengimport data Excel.')
            return
        file_path, _ = QFileDialog.getOpenFileName(self, 'Pilih File Excel', '', 'Excel Files (*.xlsx)')
        if not file_path:
            return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                QMessageBox.warning(self, 'Import Excel', 'File kosong atau format salah.')
                return
            headers = [str(h).strip().lower() for h in rows[0]]
            data_rows = rows[1:]
            conn = get_connection()
            c = conn.cursor()
            imported = 0
            failed = 0
            for idx, row in enumerate(data_rows, start=2):  # start=2 karena baris 1 header
                if len(row) < 3:
                    print(f"[IMPORT FAIL] Baris {idx}: tidak lengkap: {row}")
                    failed += 1
                    continue
                partnumber, nama, stok = row[:3]
                if not partnumber or not nama:
                    print(f"[IMPORT FAIL] Baris {idx}: Partnumber/Nama kosong: {row}")
                    failed += 1
                    continue
                try:
                    if stok is None or hasattr(stok, 'isoformat'):
                        stok = 0
                    else:
                        stok = int(stok)
                except Exception:
                    stok = 0
                try:
                    c.execute('SELECT id FROM barang WHERE partnumber=?', (partnumber,))
                    exists = c.fetchone()
                    if exists:
                        c.execute('UPDATE barang SET nama=?, stok=? WHERE id=?', (nama, stok, exists[0]))
                    else:
                        c.execute('INSERT INTO barang (partnumber, nama, stok) VALUES (?, ?, ?)', (partnumber, nama, stok))
                    imported += 1
                except Exception as e:
                    print(f"[IMPORT FAIL] Baris {idx}: {partnumber} - {nama}: {e}")
                    failed += 1
            conn.commit()
            conn.close()
            self.load_data()
            QMessageBox.information(self, 'Import Excel', f'Import selesai. {imported} data berhasil diimpor, {failed} gagal. Lihat terminal untuk detail log.')
        except Exception as e:
            print(f"[IMPORT FAIL] Gagal import file: {e}")
            QMessageBox.warning(self, 'Import Excel', f'Gagal import: {e}')

    def import_data_master(self):
        from PyQt6.QtWidgets import QDialog, QVBoxLayout, QLabel, QPushButton, QTableWidget, QTableWidgetItem, QComboBox, QFileDialog, QMessageBox, QHBoxLayout
        import openpyxl
        file_path, _ = QFileDialog.getOpenFileName(self, 'Pilih File Excel Data Master', '', 'Excel Files (*.xlsx)')
        if not file_path:
            return
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                QMessageBox.warning(self, 'Import Data Master', 'File kosong atau format salah.')
                return
            headers = [str(h).strip().lower() for h in rows[0]]
            data_rows = rows[1:]
            # Mapping otomatis
            col_map = {'partnumber': -1, 'nama': -1, 'stok': -1}
            for idx, h in enumerate(headers):
                if 'part' in h and 'number' in h:
                    col_map['partnumber'] = idx
                elif 'nama' in h:
                    col_map['nama'] = idx
                elif 'stok' in h:
                    col_map['stok'] = idx
            # Jika mapping gagal, minta user pilih manual
            if col_map['partnumber'] == -1 or col_map['nama'] == -1:
                dlg = QDialog(self)
                dlg.setWindowTitle('Mapping Kolom Data Master')
                vbox = QVBoxLayout()
                vbox.addWidget(QLabel('Pilih kolom untuk Part Number, Nama Barang, dan Stok:'))
                combo_part = QComboBox(); combo_nama = QComboBox(); combo_stok = QComboBox()
                for h in headers:
                    combo_part.addItem(h)
                    combo_nama.addItem(h)
                    combo_stok.addItem(h)
                vbox.addWidget(QLabel('Part Number:'))
                vbox.addWidget(combo_part)
                vbox.addWidget(QLabel('Nama Barang:'))
                vbox.addWidget(combo_nama)
                vbox.addWidget(QLabel('Stok:'))
                vbox.addWidget(combo_stok)
                btn_ok = QPushButton('OK')
                btn_ok.clicked.connect(dlg.accept)
                vbox.addWidget(btn_ok)
                dlg.setLayout(vbox)
                if not dlg.exec():
                    return
                col_map['partnumber'] = combo_part.currentIndex()
                col_map['nama'] = combo_nama.currentIndex()
                col_map['stok'] = combo_stok.currentIndex()
            # Preview data
            preview_dlg = QDialog(self)
            preview_dlg.setWindowTitle('Preview Data Master')
            vbox = QVBoxLayout()
            table = QTableWidget()
            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(['Part Number', 'Nama Barang', 'Stok'])
            table.setRowCount(len(data_rows))
            for i, row in enumerate(data_rows):
                table.setItem(i, 0, QTableWidgetItem(str(row[col_map['partnumber']]) if row[col_map['partnumber']] is not None else ''))
                table.setItem(i, 1, QTableWidgetItem(str(row[col_map['nama']]) if row[col_map['nama']] is not None else ''))
                table.setItem(i, 2, QTableWidgetItem(str(row[col_map['stok']]) if row[col_map['stok']] is not None else '0'))
            vbox.addWidget(QLabel('Preview data yang akan diimport:'))
            vbox.addWidget(table)
            btn_import = QPushButton('Import')
            btn_import.clicked.connect(preview_dlg.accept)
            vbox.addWidget(btn_import)
            preview_dlg.setLayout(vbox)
            if not preview_dlg.exec():
                return
            # Import data
            conn = get_connection()
            c = conn.cursor()
            imported = 0
            failed = 0
            for row in data_rows:
                partnumber = str(row[col_map['partnumber']]).strip() if row[col_map['partnumber']] is not None else ''
                nama = str(row[col_map['nama']]).strip() if row[col_map['nama']] is not None else ''
                try:
                    stok = int(row[col_map['stok']]) if row[col_map['stok']] is not None else 0
                except Exception:
                    stok = 0
                if not partnumber or not nama:
                    failed += 1
                    continue
                # Upsert: jika partnumber sudah ada, update; jika belum, insert
                c.execute('SELECT id FROM barang WHERE partnumber=?', (partnumber,))
                exists = c.fetchone()
                if exists:
                    c.execute('UPDATE barang SET nama=?, stok=? WHERE id=?', (nama, stok, exists[0]))
                else:
                    c.execute('INSERT INTO barang (partnumber, nama, stok) VALUES (?, ?, ?)', (partnumber, nama, stok))
                imported += 1
            conn.commit()
            conn.close()
            self.load_data()
            QMessageBox.information(self, 'Import Data Master', f'Import selesai. {imported} data berhasil diimpor, {failed} gagal.')
        except Exception as e:
            QMessageBox.warning(self, 'Import Data Master', f'Gagal import: {e}')

    # Semua fungsi barcode dihapus

    def export_list_barang_pdf(self):
        """Export seluruh list barang ke PDF landscape"""
        try:
            from utils.export_pdf import export_barang_pdf
            from PyQt6.QtWidgets import QFileDialog
            conn = get_connection()
            c = conn.cursor()
            c.execute("SELECT partnumber, nama, kategori, satuan, stok, rop FROM barang ORDER BY nama ASC")
            rows = c.fetchall()
            conn.close()

            data = []
            for row in rows:
                data.append({
                    "partnumber": row[0] or "",
                    "nama": row[1] or "",
                    "kategori": row[2] or "",
                    "satuan": row[3] or "",
                    "stok": row[4] if row[4] is not None else 0,
                    "rop": row[5] if row[5] is not None else 0,
                })

            if not data:
                QMessageBox.information(self, 'Export PDF', 'Tidak ada data barang untuk diexport.')
                return

            default_name = 'List_Barang.pdf'
            filename, _ = QFileDialog.getSaveFileName(self, 'Simpan PDF', default_name, 'PDF Files (*.pdf)')
            if not filename:
                return

            export_barang_pdf(filename, data)
            QMessageBox.information(self, 'Export PDF', f'List barang berhasil diexport: {filename}')
        except Exception as e:
            QMessageBox.warning(self, 'Export PDF', f'Gagal export list barang: {e}')

    def set_rop(self):
        """Set ROP untuk barang yang dipilih"""
        if not is_admin():
            QMessageBox.warning(self, 'Akses Ditolak', 'Hanya admin yang dapat mengatur ROP.')
            return
            
        selected = self.table.currentRow()
        if selected < 0:
            QMessageBox.warning(self, 'Pilih Data', 'Pilih baris yang akan set ROP!')
            return
        partnumber_item = self.table.item(selected, 1)
        partnumber = partnumber_item.text() if partnumber_item is not None else ''
        if not partnumber:
            QMessageBox.warning(self, 'Error', 'Part Number tidak valid!')
            return
        
        # Ambil data barang
        conn = get_connection()
        c = conn.cursor()
        c.execute("SELECT * FROM barang WHERE partnumber=?", (partnumber,))
        row = c.fetchone()
        conn.close()
        
        if not row:
            QMessageBox.warning(self, 'Error', 'Data tidak ditemukan!')
            return
        
        # Dialog untuk set ROP
        dialog = ROPDialog(self, row)
        if dialog.exec():
            rop_value = dialog.get_rop_value()
            conn = get_connection()
            c = conn.cursor()
            c.execute("UPDATE barang SET rop=? WHERE partnumber=?", (rop_value, partnumber))
            conn.commit()
            conn.close()
            user = get_current_user()
            if user:
                log_aktivitas(user['id'], 'set_rop', 'barang', row[0], f"Set ROP untuk {partnumber}: {rop_value}")
            self.load_data()
            QMessageBox.information(self, 'Set ROP', f'ROP berhasil diset untuk {partnumber}')
            try:
                self.data_changed.emit()
            except Exception:
                pass

    def open_in_floating_window(self):
        """Buka List Barang di floating window"""
        try:
            from utils.floating_window import floating_manager
            from screens.list_barang import ListBarangScreen
            floating_widget = ListBarangScreen(None, floating_mode=True)
            floating_window = floating_manager.create_floating_window(
                f"list_barang_{id(floating_widget)}",
                "List Barang",
                floating_widget
            )
            floating_window.show_floating()
        except Exception as e:
            QMessageBox.warning(self, 'Floating Window', f'Error: {e}')

    def save_all_barang(self):
        # Placeholder: reload data atau commit perubahan jika ada fitur edit langsung
        self.load_data()
        QMessageBox.information(self, 'Simpan', 'Data barang berhasil disimpan/diupdate!')

class BarangDialog(QDialog):
    def __init__(self, parent=None, data=None):
        super().__init__(parent)
        self.setWindowTitle('Input Barang')
        self.setModal(True)
        self.resize(400, 250)
        layout = QFormLayout()
        self.partnumber = QLineEdit()
        self.nama = QLineEdit()
        self.satuan = QLineEdit()
        self.satuan.setPlaceholderText("pcs / liter / botol ...")
        layout.addRow('Satuan:', self.satuan)
        self.kategori = QComboBox()
        # Kategori hanya peruntukan
        categories = ['project', 'retail', 'stock', 'warranty']
        self.kategori.addItems(categories)
        self.kategori.setEditable(True)
        self.kategori.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.kategori.setToolTip('Pilih atau ketik kategori baru')
        self.kategori.setStyleSheet('''
            QComboBox {
                padding: 8px;
                border-radius: 8px;
                border: 1px solid #ccc;
                font-size: 14px;
                background: white;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 5px solid #666;
                margin-right: 5px;
            }
            QComboBox QAbstractItemView {
                border: 1px solid #ccc;
                border-radius: 8px;
                background: white;
                selection-background-color: #FF2800;
                selection-color: white;
                outline: none;
                max-height: 200px;
            }
            QComboBox QAbstractItemView::item {
                padding: 8px 12px;
                border-bottom: 1px solid #f0f0f0;
            }
            QComboBox QAbstractItemView::item:hover {
                background-color: #FFF3F0;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #FF2800;
                color: white;
            }
            QComboBox QAbstractItemView QScrollBar:vertical {
                background: #F0F0F0;
                width: 8px;
                border-radius: 4px;
            }
            QComboBox QAbstractItemView QScrollBar::handle:vertical {
                background: #C0C0C0;
                border-radius: 4px;
                min-height: 20px;
            }
            QComboBox QAbstractItemView QScrollBar::handle:vertical:hover {
                background: #A0A0A0;
            }
            QComboBox QAbstractItemView QScrollBar::add-line:vertical, 
            QComboBox QAbstractItemView QScrollBar::sub-line:vertical {
                height: 0px;
            }
        ''')
        self.rop = QSpinBox()
        self.rop.setRange(0, 100000)
        self.rop.setToolTip('Reorder Point - Batas minimum stok sebelum perlu restock')
        layout.addRow('Part Number:', self.partnumber)
        layout.addRow('Nama Barang:', self.nama)
        layout.addRow('Kategori:', self.kategori)
        layout.addRow('ROP (Reorder Point):', self.rop)
        # Tombol Simpan & Batal manual
        btn_row = QHBoxLayout()
        self.btn_save = QPushButton('Simpan')
        self.btn_save.setStyleSheet('background:#FF9800; color:#FFF; font-size:18px; padding:8px 32px; border-radius:10px; font-weight:bold;')
        self.btn_save.clicked.connect(self.accept)
        self.btn_cancel = QPushButton('Batal')
        self.btn_cancel.setStyleSheet('background:#FFF; color:#FF2800; font-size:18px; padding:8px 32px; border:2px solid #FF2800; border-radius:10px; font-weight:bold;')
        self.btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_save)
        btn_row.addWidget(self.btn_cancel)
        layout.addRow(btn_row)
        self.setLayout(layout)
        # Prefill jika edit
        if data:
            self.partnumber.setText(str(data[1]) if data[1] is not None else '')
            self.nama.setText(str(data[2]) if data[2] is not None else '')
            # Set kategori value
            try:
                kategori_val = data[3] if len(data) > 3 else ''  # kategori is at index 3
                if kategori_val:
                    index = self.kategori.findText(str(kategori_val))
                    if index >= 0:
                        self.kategori.setCurrentIndex(index)
                    else:
                        self.kategori.setEditText(str(kategori_val))
            except Exception:
                pass
            # Set ROP value
            try:
                rop_val = data[6] if len(data) > 6 else 0  # rop is at index 6
                if rop_val is None:
                    self.rop.setValue(0)
                else:
                    self.rop.setValue(int(rop_val))
            except Exception:
                self.rop.setValue(0)
            # set satuan value
            try :
                satuan_val = data[7] if len(data) > 7 else ''
                self.satuan.setText(str(satuan_val) if satuan_val else '')
            except exception:
                self.satuan.setText('')
        # Validasi input
        self.partnumber.setMaxLength(32)
        self.nama.setMaxLength(100)
        self.partnumber.setPlaceholderText('Wajib diisi')
        self.nama.setPlaceholderText('Wajib diisi')
    def accept(self):
        if not self.partnumber.text().strip() or not self.nama.text().strip():
            QMessageBox.warning(self, 'Validasi', 'Part Number dan Nama Barang wajib diisi!')
            return
        super().accept()
    def get_data(self):
        partnumber = self.partnumber.text().strip()
        nama = self.nama.text().strip()
        kategori = self.kategori.currentText().strip()
        satuan = self.satuan.text().strip()
        # Stok selalu 0 untuk barang baru, untuk edit akan dipertahankan di database
        stok = 0
        rop = self.rop.value()
        
        return {
            'partnumber': partnumber,
            'nama': nama,
            'kategori': kategori,
            'stok': stok,
            'rop': rop,
            'satuan': satuan,
        }

class ROPDialog(QDialog):
    def __init__(self, parent=None, data=None):
        super().__init__(parent)
        self.setWindowTitle('Set ROP (Reorder Point)')
        self.setModal(True)
        self.resize(400, 200)
        self.data = data
        
        layout = QVBoxLayout()
        
        # Info barang
        if data:
            info_label = QLabel(f'Part Number: {data[1]}\nNama: {data[2]}\nStok Saat Ini: {data[4]}')
            info_label.setStyleSheet('font-size: 16px; margin-bottom: 16px;')
            layout.addWidget(info_label)
        
        # ROP input
        rop_layout = QHBoxLayout()
        rop_layout.addWidget(QLabel('ROP (Reorder Point):'))
        self.rop_input = QSpinBox()
        self.rop_input.setRange(0, 100000)
        self.rop_input.setValue(data[6] if data and len(data) > 6 else 0)
        self.rop_input.setToolTip('Batas minimum stok sebelum perlu restock')
        rop_layout.addWidget(self.rop_input)
        layout.addLayout(rop_layout)
        
        # Tombol
        btn_layout = QHBoxLayout()
        self.btn_save = QPushButton('Simpan')
        self.btn_save.setStyleSheet('background:#FF9800; color:#FFF; font-size:16px; padding:8px 24px; border-radius:8px; font-weight:bold;')
        self.btn_save.clicked.connect(self.accept)
        self.btn_cancel = QPushButton('Batal')
        self.btn_cancel.setStyleSheet('background:#FFF; color:#FF2800; font-size:16px; padding:8px 24px; border:2px solid #FF2800; border-radius:8px; font-weight:bold;')
        self.btn_cancel.clicked.connect(self.reject)
        btn_layout.addWidget(self.btn_save)
        btn_layout.addWidget(self.btn_cancel)
        layout.addLayout(btn_layout)
        
        self.setLayout(layout)
    
    def get_rop_value(self):
        return self.rop_input.value() 